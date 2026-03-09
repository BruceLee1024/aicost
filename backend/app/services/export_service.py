"""Export service: generate Excel valuation reports."""

from __future__ import annotations

import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.boq_item import BoqItem
from app.models.calc_result import CalcResult
from app.models.line_item_quota_binding import LineItemQuotaBinding
from app.models.project import Project
from app.models.quota_item import QuotaItem
from app.services.pricing_engine import DEFAULT_FEE_CONFIG
from app.services.project_calc_service import run_project_calculation
from app.services.snapshot_service import diff_snapshots


def export_valuation_report(project_id: int, db: Session) -> bytes:
    """Generate an Excel valuation report and return as bytes."""

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise ValueError(f"Project {project_id} not found")

    summary, line_results = run_project_calculation(project_id=project_id, db=db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分部分项组价表"

    # --- Styles ---
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # --- Title ---
    ws.merge_cells("A1:L1")
    ws["A1"] = f"分部分项组价表 — {project.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:L2")
    ws["A2"] = f"生成时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  地区: {project.region}"
    ws["A2"].font = Font(size=9, italic=True)

    # --- Header row ---
    headers = [
        "序号", "编码", "名称", "单位", "工程量",
        "人工费", "材料费", "机械费", "直接费",
        "管理费", "利润", "规费", "税前合计", "税金", "合计",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # --- Data rows ---
    for row_idx, (boq, result) in enumerate(line_results, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4)
        ws.cell(row=row_idx, column=2, value=boq.code)
        ws.cell(row=row_idx, column=3, value=boq.name)
        ws.cell(row=row_idx, column=4, value=boq.unit)
        ws.cell(row=row_idx, column=5, value=boq.quantity)
        ws.cell(row=row_idx, column=6, value=result.labor_cost)
        ws.cell(row=row_idx, column=7, value=result.material_cost)
        ws.cell(row=row_idx, column=8, value=result.machine_cost)
        ws.cell(row=row_idx, column=9, value=result.direct_cost)
        ws.cell(row=row_idx, column=10, value=result.management_fee)
        ws.cell(row=row_idx, column=11, value=result.profit)
        ws.cell(row=row_idx, column=12, value=result.regulatory_fee)
        ws.cell(row=row_idx, column=13, value=result.pre_tax_total)
        ws.cell(row=row_idx, column=14, value=result.tax)
        ws.cell(row=row_idx, column=15, value=result.total)

    # --- Summary row ---
    sum_row = len(line_results) + 5
    ws.cell(row=sum_row, column=1, value="合计").font = header_font
    ws.cell(row=sum_row, column=9, value=summary.total_direct).font = header_font
    ws.cell(row=sum_row, column=10, value=summary.total_management).font = header_font
    ws.cell(row=sum_row, column=11, value=summary.total_profit).font = header_font
    ws.cell(row=sum_row, column=12, value=summary.total_regulatory).font = header_font
    ws.cell(row=sum_row, column=13, value=summary.total_pre_tax).font = header_font
    ws.cell(row=sum_row, column=14, value=summary.total_tax).font = header_font
    ws.cell(row=sum_row, column=15, value=summary.grand_total).font = header_font

    # --- Column widths ---
    col_widths = [6, 12, 20, 6, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Division summary sheet ---
    from collections import defaultdict
    div_totals: dict[str, float] = defaultdict(float)
    for boq, result in line_results:
        div_name = boq.division or "未分类"
        div_totals[div_name] += result.total

    if div_totals:
        ws2 = wb.create_sheet(title="分部汇总")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = f"分部汇总表 — {project.name}"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2["A1"].alignment = center

        for col_idx, h in enumerate(["分部", "合计金额", "占比"], 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        grand = summary.grand_total or 1
        for row_idx, (div, total) in enumerate(sorted(div_totals.items()), 4):
            ws2.cell(row=row_idx, column=1, value=div)
            ws2.cell(row=row_idx, column=2, value=round(total, 2))
            ws2.cell(row=row_idx, column=3, value=f"{round(total / grand * 100, 1)}%")

        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_diff_report(snap_a, snap_b) -> bytes:
    """Generate an Excel diff report comparing two snapshots."""
    from app.models.snapshot import Snapshot  # avoid circular import

    report = diff_snapshots(snap_a, snap_b)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异对比表"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "版本差异对比表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"快照 A (ID {report.snapshot_a_id}) vs 快照 B (ID {report.snapshot_b_id})  |  "
        f"生成时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A2"].font = Font(size=9, italic=True)

    # Summary
    ws.cell(row=3, column=1, value="旧合计").font = header_font
    ws.cell(row=3, column=2, value=report.old_grand_total)
    ws.cell(row=3, column=3, value="新合计").font = header_font
    ws.cell(row=3, column=4, value=report.new_grand_total)
    ws.cell(row=3, column=5, value="差额").font = header_font
    ws.cell(row=3, column=6, value=report.grand_total_delta)

    # Headers
    headers = ["编码", "名称", "变更类型", "旧金额", "新金额", "差额"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, line in enumerate(report.lines, 6):
        ws.cell(row=row_idx, column=1, value=line.boq_code)
        ws.cell(row=row_idx, column=2, value=line.boq_name)
        ws.cell(row=row_idx, column=3, value=line.change_type)
        ws.cell(row=row_idx, column=4, value=line.old_total)
        ws.cell(row=row_idx, column=5, value=line.new_total)
        ws.cell(row=row_idx, column=6, value=line.delta)

    col_widths = [12, 20, 10, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
